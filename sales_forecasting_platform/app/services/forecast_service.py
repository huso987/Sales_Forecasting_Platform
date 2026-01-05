# app/services/forecast_service.py

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.services.data_service import DataService
from app.services.training_service import TrainingService
from app.core.mailer import Mailer

class ForecastService:

    def run(self, urunler, models, months, holdout, email):

        data = DataService().load()
        trainer = TrainingService()

        products = [u.strip() for u in urunler.split(",")]
        selected_models = models

        all_rows = []

        for product in products:
            # 🔥 BURASI DÜZELTİLDİ
            series = data[data["URUN_KODU"] == product]["MIKTAR"]

            if len(series) <= holdout + 2:
                continue

            results = trainer.train_and_evaluate(
                series=series,
                selected_models=selected_models,
                holdout=holdout
            )

            if not results:
                continue

            best_model_name = min(
                results,
                key=lambda k: results[k]["mape"]
            )

            for model_name, info in results.items():
                forecast = info["model"].predict(months)

                for i, val in enumerate(forecast):
                    all_rows.append({
                        "URUN": product,
                        "MODEL": model_name,
                        "AY": i + 1,
                        "TAHMIN": round(float(val), 2),
                        "MAPE": round(info["mape"], 4),
                        "BEST": model_name == best_model_name
                    })

        df = pd.DataFrame(all_rows)

        output_path = "outputs/forecast_results.xlsx"
        df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            if ws[f"F{row}"].value is True:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = yellow

        wb.save(output_path)

        Mailer().send(email, output_path)
